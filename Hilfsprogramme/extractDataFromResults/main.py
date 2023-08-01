import os
import re
import shutil
import sys

from session_scrapper import Session
from argparser_scrapper import Argparser
from runtime_scrapper import Runtime
from image_scrapper import Image
from execution_check import ExecutionCheck
from profiler_scrapper import Profiler
from monster_runtime import MonsterRuntime
from monster_profiler import MonsterProfiler
from openpyxl import Workbook, load_workbook
import os


def save_to_excel_seperat(datum, uhrzeit, benchmark_path, argparser, session, runtimes, image, profilers, execution_check, filename):

    # Laden oder erstellen der Excel-Datei
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()



    # Behandlung des Argparser-Objekts
    if argparser is not None:
        if "Argparser" in wb.sheetnames:
            ws = wb["Argparser"]
            start_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title="Argparser")
            attrs = vars(argparser)
            for col, attr in enumerate(attrs, start=1):
                ws.cell(row=1, column=col, value=attr)
            start_row = 2

        attrs = vars(argparser)
        for col, attr in enumerate(attrs, start=1):
            ws.cell(row=start_row, column=col, value=getattr(argparser, attr))

    # Behandlung des Session-Objekts
    if session is not None:
        if "Session" in wb.sheetnames:
            ws = wb["Session"]
            start_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title="Session")
            attrs = vars(session)
            for col, attr in enumerate(attrs, start=1):
                ws.cell(row=1, column=col, value=attr)
            start_row = 2

        attrs = vars(session)
        for col, attr in enumerate(attrs, start=1):
            ws.cell(row=start_row, column=col, value=getattr(session, attr))

    # Behandlung des Image-Objekts
    if image is not None:
        if "Image" in wb.sheetnames:
            ws = wb["Image"]
            start_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title="Image")
            attrs = vars(image)
            for col, attr in enumerate(attrs, start=1):
                ws.cell(row=1, column=col, value=attr)
            start_row = 2

        attrs = vars(image)
        for col, attr in enumerate(attrs, start=1):
            ws.cell(row=start_row, column=col, value=getattr(image, attr))

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Speichern der Excel-Datei
    wb.save(filename)

def save_to_excel(datum, uhrzeit, benchmark_path, ben_MapWidth, ben_MapHeight, isEmpty, argparser, session, monsterRuntime, image, monsterProfiler, execution_check, filename):
        # Laden oder erstellen der Excel-Datei
        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()

        # Einzelnes Arbeitsblatt für alle Daten
        if "Data" in wb.sheetnames:
            ws = wb["Data"]
            start_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title="Data")
            start_row = 2

        # Listen der Objekte und ihrer Namen
        objects = [argparser, session, image, execution_check, monsterRuntime, monsterProfiler]
        names = ["Argparser", "Session", "Image", "ExecutionCheck", "RunTime", "Profiler"]

        # Dummy-Objekte für den Fall, dass die ursprünglichen Objekte leer sind
        dummy_argparser = Argparser() if argparser is None else None
        dummy_session = Session() if session is None else None
        dummy_image = Image() if image is None else None
        dummy_execution_check = ExecutionCheck() if execution_check is None else None
        dummy_monsterRuntime = MonsterRuntime() if monsterRuntime is None else None
        dummy_monsterProfiler = MonsterProfiler() if monsterProfiler is None else None
        dummy_objects = [dummy_argparser, dummy_session, dummy_image, dummy_execution_check, dummy_monsterRuntime, dummy_monsterProfiler]

        # Initialisierung des Spaltenzählers
        start_col = 1

        # Hinzufügen der einzelnen Daten vor den Objektattributen
        data = [datum, uhrzeit, benchmark_path, ben_MapWidth, ben_MapHeight, isEmpty]
        data_names = ["Datum", "Uhrzeit", "benchmark_path", "Spielfeld_width", "Spielfeld_height", "isEmpty"]
        for val, name in zip(data, data_names):
            if start_row == 2:  # Header nur für die erste Zeile hinzufügen
                ws.cell(row=1, column=start_col, value=name)
            ws.cell(row=start_row, column=start_col, value=val)
            start_col += 1

        # Durchgehen jedes Objekts und Schreiben seiner Attribute in die Excel-Datei
        for obj, dummy_obj, name in zip(objects, dummy_objects, names):
            attrs = vars(obj) if obj is not None else vars(dummy_obj)
            if start_row == 2:  # Header nur für die erste Zeile hinzufügen
                for col, attr in enumerate(attrs, start=start_col):
                    ws.cell(row=1, column=col, value=f"{name}_{attr}")

            for col, attr in enumerate(attrs, start=start_col):
                ws.cell(row=start_row, column=col, value=getattr(obj, attr) if obj is not None else None)

            # Aktualisierung des Spaltenstarts für das nächste Objekt
            start_col += len(attrs)

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Speichern der Excel-Datei
        wb.save(filename)


def handle_file(filename):
    # Check if the file exists
    if os.path.exists(filename):
        # If it does, delete it
        os.remove(filename)
        print(f"{filename} wurde geloescht.")

    # Recreate the file
    open(filename, 'w').close()
    print(f"{filename} wurde neu erstellt.")


def after_main(path_to_benchmark, copy_files):
    # Sort the contents of the result files
    for filename in ['missing-results.txt', 'existing-results.txt']:
        with open(filename, 'r') as f:
            lines = f.readlines()
        lines.sort()
        with open(filename, 'w') as f:
            f.writelines(lines)

    # handle_file('missing-results.txt')
    # with open('existing-results.txt', 'r') as f:
    #     lines = f.readlines()
    # lines.sort()
    # with open('existing-results.txt', 'w') as f:
    #     f.writelines(lines)

    # If copy_files is set to True, copy the missing files to the new location
    if copy_files:
        if os.path.exists('./benchmark'):
            shutil.rmtree('./benchmark')
        os.makedirs('./benchmark', exist_ok=True)

        with open('missing-results.txt', 'r') as f:
            missing_files = f.readlines()

        for file in missing_files:
            file = file.strip()
            new_file_path = file.replace('benchmark', path_to_benchmark, 1)
            local_file_path = file.replace('benchmark', './benchmark', 1)
            local_dir_path = os.path.dirname(local_file_path)

            os.makedirs(local_dir_path, exist_ok=True)

            if os.path.exists(new_file_path):
                shutil.copy2(new_file_path, local_file_path)

def format_time(datum, uhrzeit):
    tag, monat, jahr = datum.split('-')
    stunden, minuten = uhrzeit.split('-')
    formatiertes_datum = f'{tag}.{monat}.{jahr}'
    formatierte_uhrzeit = f'{stunden}:{minuten}'
    return formatiertes_datum, formatierte_uhrzeit

def get_benchmark_path(filepath):
    # Extrahiere das Datum und die Uhrzeit
    match = re.match(r"\.\/logs\/(\d{2}-\d{2}-\d{4})--(\d{2}-\d{2})-(.+\.log)", filepath)
    if not match:
        raise ValueError(f"Das Format von '{filepath}' wird nicht erkannt.")

    datum, uhrzeit, rest = match.groups()

    print(f"Originaler Dateipfad: {filepath}")
    print(f"Datum: {datum}")
    print(f"Uhrzeit: {uhrzeit}")

    # Ersetze '__' durch '/' und entferne '.log'
    rest_anpassung = rest.replace("__", "/")[:-4]
    print(f"Rest nach Ersetzung von '__' durch '/': {rest_anpassung}")

    # Überprüfe, ob das viertletzte Zeichen '-' ist und ersetze es ggf. durch '.'
    if rest_anpassung[-4] == '-':
        benchmark_path = rest_anpassung[:-4] + '.' + rest_anpassung[-3:]
    else:
        benchmark_path = rest_anpassung

    print(f"Benchmark-Pfad: {benchmark_path}")

    datum, uhrzeit = format_time(datum, uhrzeit)

    return datum, uhrzeit, benchmark_path

def benchmark_parameter(path_to_benchmark, benchmark_path):
    ben_MapWidth = -1
    ben_MapHeight = -1

    # Erstes Vorkommen von 'benchmark' entfernen
    benchmark_path = benchmark_path.replace('benchmark/', '', 1)

    # Zusammenführen von path_to_benchmark und benchmark_path
    combined_path = os.path.join(path_to_benchmark, benchmark_path)

    # Ausgabe des Pfades
    print(f'Combined path: {combined_path}')

    # Überprüfen, ob die Datei existiert und Ausgabe des Ergebnisses
    if os.path.isfile(combined_path):
        print(f'The file {combined_path} exists.')
        ben_MapWidth, ben_MapHeight = read_file_dimensions(combined_path)
    else:
        print(f'The file {combined_path} does not exist.')

    return ben_MapWidth, ben_MapHeight

def read_file_dimensions(filepath):
    with open(filepath, 'r') as file:
        file.readline()  # überspringt die erste Zeile
        dimensions = file.readline()  # liest die zweite Zeile
        width, height = map(int, dimensions.split())  # teilt die Zeile auf und konvertiert in Zahlen

    return width, height

def main(path_to_benchmark):
    logs_folder = "./logs"
    for file in os.listdir(logs_folder):
        if file.endswith(".log"):
            isEmpty = False
            argparser = None
            session = None
            runtimes = None
            image = None
            profilers = None
            monsterRuntime = None
            execution_check = None

            filepath = os.path.join(logs_folder, file)
            datum, uhrzeit, benchmark_path = get_benchmark_path(filepath)

            if "benchmark_2/" in benchmark_path:
                benchmark_path = benchmark_path.replace("benchmark_2/", "benchmark/")

            ben_MapWidth, ben_MapHeight = benchmark_parameter(path_to_benchmark, benchmark_path)



            if os.stat(filepath).st_size == 0:
                isEmpty = True
                print(f"Die Datei {filepath} ist leer. Überspringe diese Datei.")
                with open('missing-results.txt', 'a') as f:
                    f.write(f"{benchmark_path}\n")

            if not isEmpty:
                argparser = Argparser.process(filepath)
                session = Session.process(filepath)
                runtimes = Runtime.process(filepath)
                image = Image.process(filepath)
                profilers = Profiler.process(filepath)

                if argparser:
                    execution_check = ExecutionCheck.process(filepath, argparser, image)

                if session:
                    print(session)
                if argparser:
                    print(argparser)
                if runtimes:
                    for runtime in runtimes:
                        print(runtime)
                    monsterRuntime = MonsterRuntime(runtimes)
                if image:
                    print(image)
                if profilers:
                    merged_profilers = profilers[0] + profilers[1]
                    for profiler in merged_profilers:
                        print(profiler)
                    monsterProfiler = MonsterProfiler(merged_profilers)

                if execution_check:
                    print(execution_check)

            save_to_excel(datum, uhrzeit, benchmark_path, ben_MapWidth, ben_MapHeight, isEmpty, argparser, session, monsterRuntime, image, monsterProfiler, execution_check, "output.xlsx")


if __name__ == "__main__":
    copy_files = True  # Set this to False if you do not want to copy the files
    path_to_benchmark = "/Users/mama/Documents/GitHub/all_benchmarks"
    handle_file('missing-results_2.txt')
    main(path_to_benchmark)
    after_main(path_to_benchmark, copy_files)


